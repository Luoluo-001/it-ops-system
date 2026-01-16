from flask import Flask, request, jsonify, send_from_directory, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from datetime import datetime, timedelta
from sqlalchemy import inspect, text
import io
import os

import threading
import json
import urllib.request
from werkzeug.utils import secure_filename
try:
    from croniter import croniter
except ImportError:
    croniter = None
try:
    from dotenv import load_dotenv

    # 加载环境变量
    load_dotenv()
except ImportError:
    print("Warning: python-dotenv not installed, skipping .env loading")



from werkzeug.security import generate_password_hash, check_password_hash
import secrets


app = Flask(__name__, static_folder='static', static_url_path='')


# 配置密钥（固定密钥，避免重启后session失效）
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'flzx-it-ops-system-2024-secret-key-do-not-change')

# Session配置（简化配置，使用默认值）
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24小时

# CORS配置 - 本地开发不需要CORS
# CORS(app, supports_credentials=True)

# 配置数据库
# 优先从环境变量读取数据库URI，支持MySQL和SQLite
# MySQL格式: mysql+pymysql://user:password@host:port/dbname
# SQLite格式: sqlite:///it_ops.db
default_db_path = os.path.join(app.root_path, 'instance', 'it_ops.db')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URI', f'sqlite:///{default_db_path}')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)

# ==================== 数据模型 ====================

class BusinessSystem(db.Model):
    """业务系统表"""
    id = db.Column(db.Integer, primary_key=True)
    system_name = db.Column(db.String(100), nullable=False, unique=True)
    system_code = db.Column(db.String(50), unique=True)
    database = db.Column(db.String(100))  # 数据库类型
    database_version = db.Column(db.String(50))  # 数据库版本
    department = db.Column(db.String(100))  # 管理部室
    department_status = db.Column(db.String(50))  # 部室状态
    status = db.Column(db.String(20), default='运行中')  # 系统状态
    description = db.Column(db.Text)  # 系统描述
    contact_person = db.Column(db.String(50))  # 负责人
    contact_phone = db.Column(db.String(20))  # 联系电话
    contact_email = db.Column(db.String(100))  # 联系邮箱
    construction_unit = db.Column(db.String(200))  # 建设单位
    location = db.Column(db.String(200))  # 所在位置
    access_url = db.Column(db.String(500))  # 访问地址
    created_at = db.Column(db.DateTime, default=datetime.now)

    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    
    events = db.relationship('Event', backref='business_system', lazy=True)
    hosts = db.relationship('SystemHost', backref='business_system', lazy=True, cascade='all, delete-orphan')
    middlewares = db.relationship('SystemMiddleware', backref='business_system', lazy=True, cascade='all, delete-orphan')
    integrations = db.relationship('SystemIntegration', backref='business_system', lazy=True, cascade='all, delete-orphan')

class SystemHost(db.Model):

    """系统主机表"""
    id = db.Column(db.Integer, primary_key=True)
    system_id = db.Column(db.Integer, db.ForeignKey('business_system.id'), nullable=False)
    host_type = db.Column(db.String(50))  # 主机类型
    ip_address = db.Column(db.String(100))  # IP地址
    host_purpose = db.Column(db.String(200))  # 主机用途
    os_version = db.Column(db.String(100))  # 操作系统版本
    cpu_cores = db.Column(db.String(50))   # CPU核数
    memory_gb = db.Column(db.String(50))   # 内存GB
    disk_gb = db.Column(db.String(50))     # 磁盘GB
    cpu_arch = db.Column(db.String(50))    # CPU架构
    created_at = db.Column(db.DateTime, default=datetime.now)


class SystemMiddleware(db.Model):
    """系统中间件表"""
    id = db.Column(db.Integer, primary_key=True)
    system_id = db.Column(db.Integer, db.ForeignKey('business_system.id'), nullable=False)
    middleware_type = db.Column(db.String(100))  # 中间件类型
    middleware_version = db.Column(db.String(50))  # 中间件版本
    quantity = db.Column(db.Integer, default=1)  # 数量
    created_at = db.Column(db.DateTime, default=datetime.now)

class SystemIntegration(db.Model):
    """系统集成关系表"""
    id = db.Column(db.Integer, primary_key=True)
    system_id = db.Column(db.Integer, db.ForeignKey('business_system.id'), nullable=False)
    integration_type = db.Column(db.String(20))  # upstream (上游) / downstream (下游)
    remote_system_name = db.Column(db.String(100), nullable=False)  # 关联系统名称
    network_type = db.Column(db.String(20))  # internal (内网) / external (外网)
    created_at = db.Column(db.DateTime, default=datetime.now)

class Event(db.Model):

    """运维事件表"""
    id = db.Column(db.Integer, primary_key=True)
    event_no = db.Column(db.String(50), unique=True, nullable=False)  # 事件编号
    system_id = db.Column(db.Integer, db.ForeignKey('business_system.id'), nullable=False)
    system_name = db.Column(db.String(100), nullable=False)  # 冗余字段,方便查询
    event_type = db.Column(db.String(50), nullable=False)  # 事件类型
    event_category = db.Column(db.String(50))  # 故障分类
    severity = db.Column(db.String(20))  # 严重程度
    status = db.Column(db.String(20), default='处理中')  # 状态:待处理,处理中,已解决,已关闭
    title = db.Column(db.String(200), nullable=False)  # 事件标题
    description = db.Column(db.Text)  # 事件描述
    occurred_at = db.Column(db.DateTime, nullable=False)  # 发生时间
    reported_by = db.Column(db.String(50))  # 报告人
    assigned_to = db.Column(db.String(50))  # 处理人
    resolution = db.Column(db.Text)  # 解决方案
    root_cause = db.Column(db.Text)  # 根本原因
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    resolved_at = db.Column(db.DateTime)  # 解决时间
    closed_at = db.Column(db.DateTime)  # 关闭时间
    progress_status = db.Column(db.String(20), default='未解决') # 处置进度: 未解决, 已解决, 已挂起
    
    processes = db.relationship('EventProcess', backref='event', lazy=True, cascade='all, delete-orphan')

    attachments = db.relationship('EventAttachment', backref='event', lazy=True, cascade='all, delete-orphan')

class EventProcess(db.Model):
    """事件处置流程表"""
    id = db.Column(db.Integer, primary_key=True)
    event_id = db.Column(db.Integer, db.ForeignKey('event.id'), nullable=False)
    step_no = db.Column(db.Integer)  # 步骤序号
    action = db.Column(db.String(200), nullable=False)  # 处置动作
    result = db.Column(db.Text)  # 处置结果
    operator = db.Column(db.String(50))  # 操作人
    operated_at = db.Column(db.DateTime, default=datetime.now)  # 操作时间
    remarks = db.Column(db.Text)  # 备注

class EventAttachment(db.Model):
    """事件附件表"""
    id = db.Column(db.Integer, primary_key=True)
    event_id = db.Column(db.Integer, db.ForeignKey('event.id'), nullable=False)
    file_name = db.Column(db.String(200), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    file_type = db.Column(db.String(50))  # 文件类型
    file_size = db.Column(db.Integer)  # 文件大小(字节)
    uploaded_by = db.Column(db.String(50))
    uploaded_at = db.Column(db.DateTime, default=datetime.now)

class SystemConfig(db.Model):
    """系统配置表"""
    id = db.Column(db.Integer, primary_key=True)
    config_key = db.Column(db.String(100), unique=True, nullable=False)
    config_value = db.Column(db.Text)
    config_type = db.Column(db.String(50))  # 配置类型
    description = db.Column(db.String(200))
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

class User(db.Model):
    """用户表"""
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    real_name = db.Column(db.String(50), nullable=False)  # 真实姓名
    role = db.Column(db.String(20), default='user')  # 角色: admin/user
    department = db.Column(db.String(100))  # 所属部门
    phone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    is_active = db.Column(db.Boolean, default=True)  # 是否启用
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    last_login = db.Column(db.DateTime)  # 最后登录时间
    
    def set_password(self, password):
        """设置密码"""
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        """验证密码"""
        return check_password_hash(self.password_hash, password)

class PlanTask(db.Model):
    """计划任务主表"""
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    task_type = db.Column(db.String(50), default='其他')
    schedule_type = db.Column(db.String(20), default='once')  # once/daily/weekly/monthly/cron
    schedule_value = db.Column(db.String(100))  # 周期附加信息（如cron表达式）
    plan_time = db.Column(db.DateTime, nullable=False)
    reminder_minutes = db.Column(db.Integer, default=1440)  # 提前提醒的分钟数
    reminder_enabled = db.Column(db.Boolean, default=True)
    reminder_sent = db.Column(db.Boolean, default=False)
    alert_robot = db.Column(db.String(100), default='默认钉钉机器人')
    webhook_url = db.Column(db.String(500))
    reminder_message = db.Column(db.Text, default='【计划任务提醒】任务：{title}，计划时间：{plan_time}，负责人：{owner}。请提前准备：{preparations}')
    status = db.Column(db.String(20), default='待执行')  # 待执行/进行中/已完成/已取消
    responsible = db.Column(db.String(200))  # 责任人（逗号分隔）
    owner = db.Column(db.String(50))  # 主负责人
    description = db.Column(db.Text)
    result_status = db.Column(db.String(20))  # 成功/部分成功/失败
    result_notes = db.Column(db.Text)

    created_by = db.Column(db.String(50))
    actual_start = db.Column(db.DateTime)
    actual_finish = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    
    preparations = db.relationship('PlanTaskPreparation', backref='task', lazy=True, cascade='all, delete-orphan')



class PlanTaskPreparation(db.Model):
    """任务准备事项"""
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey('plan_task.id'), nullable=False)
    description = db.Column(db.String(300), nullable=False)
    status = db.Column(db.String(20), default='未开始')  # 未开始/进行中/已完成
    estimated_minutes = db.Column(db.Integer)
    order_no = db.Column(db.Integer, default=1)

class NotificationAudit(db.Model):
    """通知审计日志表"""
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey('plan_task.id'), nullable=True)
    task_title = db.Column(db.String(200))
    robot_name = db.Column(db.String(100))
    webhook_url = db.Column(db.String(500))
    msg_type = db.Column(db.String(20), default='markdown')
    title = db.Column(db.String(200))
    content = db.Column(db.Text)
    status = db.Column(db.String(20))  # 成功/失败
    error_msg = db.Column(db.Text)
    sent_at = db.Column(db.DateTime, default=datetime.now)



def ensure_plan_task_schema():
    """确保计划任务相关表和字段存在（兼容旧版SQLite数据库）。"""
    with app.app_context():
        # 确保所有模型对应的表存在
        db.create_all()

        inspector = inspect(db.engine)
        table_names = set(inspector.get_table_names())

        # 如果旧库没有表，显式创建
        if 'plan_task' not in table_names:
            PlanTask.__table__.create(db.engine, checkfirst=True)
        if 'plan_task_preparation' not in table_names:
            PlanTaskPreparation.__table__.create(db.engine, checkfirst=True)
        if 'notification_audit' not in table_names:
            NotificationAudit.__table__.create(db.engine, checkfirst=True)
        if 'system_integration' not in table_names:
            SystemIntegration.__table__.create(db.engine, checkfirst=True)



        # 补齐 plan_task 缺失字段
        inspector = inspect(db.engine)
        plan_cols = {col['name'] for col in inspector.get_columns('plan_task')}
        column_defs = {
            'schedule_type': 'VARCHAR(20)',
            'schedule_value': 'VARCHAR(100)',
            'plan_time': 'DATETIME',
            'reminder_minutes': 'INTEGER',
            'reminder_enabled': 'BOOLEAN',
            'alert_robot': 'VARCHAR(100)',
            'webhook_url': 'VARCHAR(500)',
            'reminder_message': 'TEXT',
            'result_status': 'VARCHAR(20)',
            'result_notes': 'TEXT',
            'actual_start': 'DATETIME',
            'actual_finish': 'DATETIME',
            'created_by': 'VARCHAR(50)',
            'reminder_sent': 'BOOLEAN DEFAULT 0'
        }


        for col, ddl in column_defs.items():
            if col not in plan_cols:
                with db.engine.begin() as conn:
                    conn.exec_driver_sql(f"ALTER TABLE plan_task ADD COLUMN {col} {ddl}")

        # 补齐 system_host 缺失字段
        host_cols = {col['name'] for col in inspector.get_columns('system_host')}
        host_column_defs = {
            'os_version': 'VARCHAR(100)',
            'cpu_cores': 'VARCHAR(50)',
            'memory_gb': 'VARCHAR(50)',
            'disk_gb': 'VARCHAR(50)',
            'cpu_arch': 'VARCHAR(50)'
        }
        for col, ddl in host_column_defs.items():
            if col not in host_cols:
                with db.engine.begin() as conn:
                    conn.exec_driver_sql(f"ALTER TABLE system_host ADD COLUMN {col} {ddl}")

        # 补齐 event 缺失字段

        event_cols = {col['name'] for col in inspector.get_columns('event')}
        event_column_defs = {
            'progress_status': "VARCHAR(20) DEFAULT '未解决'",
            'event_category': "VARCHAR(50)",
            'system_name': "VARCHAR(100)"
        }
        for col, ddl in event_column_defs.items():
            if col not in event_cols:
                with db.engine.begin() as conn:
                    conn.exec_driver_sql(f"ALTER TABLE event ADD COLUMN {col} {ddl}")

        # 补齐 event_process 缺失字段
        process_cols = {col['name'] for col in inspector.get_columns('event_process')}
        process_column_defs = {
            'operated_at': "DATETIME DEFAULT CURRENT_TIMESTAMP"
        }
        for col, ddl in process_column_defs.items():
            if col not in process_cols:
                with db.engine.begin() as conn:
                    conn.exec_driver_sql(f"ALTER TABLE event_process ADD COLUMN {col} {ddl}")

        # 补齐 business_system 缺失字段
        system_cols = {col['name'] for col in inspector.get_columns('business_system')}
        if 'construction_unit' not in system_cols:
            with db.engine.begin() as conn:
                conn.exec_driver_sql("ALTER TABLE business_system ADD COLUMN construction_unit VARCHAR(200)")
        
        if 'location' not in system_cols:
            with db.engine.begin() as conn:
                conn.exec_driver_sql("ALTER TABLE business_system ADD COLUMN location VARCHAR(200)")
        
        if 'access_url' not in system_cols:
            with db.engine.begin() as conn:
                conn.exec_driver_sql("ALTER TABLE business_system ADD COLUMN access_url VARCHAR(500)")





schema_bootstrapped = False
schema_bootstrap_lock = threading.Lock()

def bootstrap_schema():
    global schema_bootstrapped
    if schema_bootstrapped:
        return
    with schema_bootstrap_lock:
        if schema_bootstrapped:
            return
        ensure_plan_task_schema()
        schema_bootstrapped = True

app.before_request(bootstrap_schema)

# 模块加载时先执行一次以避免首个请求前的查询因缺列报错
try:
    bootstrap_schema()
except Exception as e:
    print(f"[schema bootstrap warning] {e}")

# 在应用启动时立即开启后台线程

def start_background_worker():
    # 避免在 Flask debug 模式下启动两次
    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true' or not app.debug:
        with app.app_context():
            reminder_thread = threading.Thread(target=background_reminder_worker, daemon=True)
            reminder_thread.start()
            print(">>> 计划任务提醒后台服务已成功启动")


def send_dingtalk_notification(webhook_url, message, title=None):


    """发送钉钉通知 (支持 Markdown 格式)"""
    if not webhook_url:
        return False, "Webhook URL为空"
    
    webhook_url = webhook_url.strip()
    
    try:
        # 如果提供了 title，说明是 Markdown 格式
        if title:
            data = {
                "msgtype": "markdown",
                "markdown": {
                    "title": title,
                    "text": message
                }
            }
        else:
            data = {
                "msgtype": "text",
                "text": {
                    "content": message
                }
            }
        
        encoded_data = json.dumps(data).encode('utf-8')
        proxy_handler = urllib.request.ProxyHandler({})
        opener = urllib.request.build_opener(proxy_handler)
        
        req = urllib.request.Request(webhook_url, data=encoded_data, headers={'Content-Type': 'application/json'})
        with opener.open(req, timeout=10) as response:
            result = response.read().decode('utf-8')
            res_json = json.loads(result)
            if res_json.get('errcode') == 0:
                return True, "发送成功"
            else:
                return False, f"钉钉返回错误: {res_json.get('errmsg')}"
    except Exception as e:
        print(f"DEBUG: 钉钉发送异常: {str(e)}")
        return False, f"发送失败: {str(e)}"




def calculate_next_run_time(current_plan_time, schedule_type, schedule_value=None, base_time=None):
    """根据周期类型计算下一次执行时间"""
    if schedule_type == 'once':
        return None
    
    if base_time is None:
        base_time = datetime.now()
    
    next_time = current_plan_time

    if schedule_type == 'daily':
        # 如果当前计划时间已经过了（或等于基准时间），加一天
        while next_time <= base_time:
            next_time += timedelta(days=1)
            
    elif schedule_type == 'weekly':
        try:
            target_weekday = int(schedule_value) # 0-6 (Mon-Sun)
        except (ValueError, TypeError):
            target_weekday = 0
            
        # 调整到目标星期
        days_ahead = target_weekday - next_time.weekday()
        if days_ahead < 0:
            days_ahead += 7
        next_time += timedelta(days=days_ahead)
        
        # 如果调整后还是过去的时间（或等于基准时间），则加一周
        while next_time <= base_time:
            next_time += timedelta(weeks=1)
            
    elif schedule_type == 'monthly':
        try:
            target_day = int(schedule_value) # 1-31
        except (ValueError, TypeError):
            target_day = 1
            
        import calendar
        # 调整到目标日期
        month = next_time.month
        year = next_time.year
        last_day = calendar.monthrange(year, month)[1]
        actual_day = min(target_day, last_day)
        next_time = next_time.replace(day=actual_day)
        
        # 如果调整后还是过去的时间（或等于基准时间），则增加月份
        while next_time <= base_time:
            month = next_time.month + 1
            year = next_time.year
            if month > 12:
                month = 1
                year += 1
            last_day = calendar.monthrange(year, month)[1]
            actual_day = min(target_day, last_day)
            next_time = next_time.replace(year=year, month=month, day=actual_day)
            
    elif schedule_type == 'cron':
        if croniter and schedule_value:
            try:
                iter = croniter(schedule_value, base_time)
                return iter.get_next(datetime)
            except Exception as e:
                print(f"DEBUG: Cron解析异常: {str(e)}")
                return next_time
        return next_time

        
    return next_time



def background_reminder_worker():
    """后台提醒检查线程"""
    import time
    with app.app_context():
        # 获取并打印服务器时间与时区偏移，辅助排查 Linux 触发问题
        import datetime as dt
        local_now = dt.datetime.now()
        utc_now = dt.datetime.utcnow()
        tz_offset = (local_now - utc_now).total_seconds() / 3600
        print(f"计划任务提醒后台线程已启动 (服务器时间: {local_now.strftime('%Y-%m-%d %H:%M:%S')}, 时区偏移: UTC{tz_offset:+.1f})")
        
        while True:
            try:
                # 显式清理 Session，确保读取最新数据库数据
                db.session.remove()
                now = datetime.now()
                # 检查所有 待执行/进行中 且开启提醒且未发送提醒的任务
                tasks = PlanTask.query.filter(
                    PlanTask.status.in_(['待执行', '进行中']),
                    PlanTask.reminder_enabled == True,
                    PlanTask.reminder_sent == False
                ).all()
                
                if tasks:
                    print(f"DEBUG: 后台扫描到 {len(tasks)} 条待提醒任务 (当前时间: {now.strftime('%H:%M:%S')})")
                
                for task in tasks:
                    current_webhook = task.webhook_url
                    
                    # 兜底：如果数据库里没存 Webhook，但存了机器人名称，尝试实时从配置中读取
                    if not current_webhook and task.alert_robot:
                        robot_config = SystemConfig.query.filter_by(config_key='alert_robots').first()
                        if robot_config:
                            try:
                                robots = json.loads(robot_config.config_value)
                                for r in robots:
                                    if r.get('name') == task.alert_robot:
                                        current_webhook = r.get('webhook')
                                        break
                            except:
                                pass

                    # 计算提醒时间点
                    reminder_time = task.plan_time - timedelta(minutes=task.reminder_minutes)
                    
                    # 打印调试日志，检查时间逻辑
                    # print(f"DEBUG: 任务[{task.title}] 计划:{task.plan_time}, 提醒:{reminder_time}, 当前:{now}")
                    
                    # 如果当前时间到达或超过提醒时间
                    if reminder_time <= now <= task.plan_time + timedelta(hours=1):
                        print(f"DEBUG: 任务[{task.title}] 满足时间条件 (提醒点:{reminder_time.strftime('%H:%M:%S')}, 计划:{task.plan_time.strftime('%H:%M:%S')})")
                        
                        if not current_webhook:
                            print(f"DEBUG: 任务[{task.title}] 满足条件但无法获取 Webhook，标记为已处理以避免死循环")
                            task.reminder_sent = True
                            db.session.commit()
                            continue

                        print(f"DEBUG: 任务[{task.title}] 准备发送通知...")
                        # 准备模板变量

                        preps = task.preparations
                        completed_count = len([p for p in preps if p.status == '已完成'])
                        prep_text_list = []
                        for p in preps:
                            status_icon = "✅" if p.status == '已完成' else "⬜"
                            prep_text_list.append(f"{status_icon} {p.description}")
                        
                        prep_text = '\n\n'.join(prep_text_list) if prep_text_list else '无'
                        prep_progress = f"{completed_count}/{len(preps)}"
                        
                        message = task.reminder_message or '任务：{title}，计划时间：{plan_time}，负责人：{owner}。'
                        # 处理换行符
                        message = message.replace('\\n', '\n').replace('\r\n', '\n')
                        
                        replacements = {
                            '{title}': task.title,
                            '{plan_time}': task.plan_time.strftime('%Y-%m-%d %H:%M'),
                            '{owner}': task.owner or '未指定',
                            '{responsible}': task.responsible or '未指定',
                            '{preparations}': prep_text,
                            '{prep_progress}': prep_progress
                        }
                        
                        for key, val in replacements.items():
                            message = message.replace(key, str(val))
                        
                        # 构建美化的 Markdown 消息
                        markdown_title = f"⏰ 计划任务提醒: {task.title}"
                        safe_message = message.replace('\n', '\n\n> ')
                        markdown_text = f"### ⏰ 计划任务提醒\n\n" \
                                        f"**任务名称**: <font color='#1d4ed8'>{task.title}</font>\n\n" \
                                        f"--- \n\n" \
                                        f"📅 **计划时间**: {task.plan_time.strftime('%Y-%m-%d %H:%M')}\n\n" \
                                        f"👤 **主负责人**: {task.owner or '未指定'}\n\n" \
                                        f"👥 **责任人**: {task.responsible or '未指定'}\n\n" \
                                        f"📊 **当前进度**: `{prep_progress}`\n\n" \
                                        f"📝 **准备事项**:\n\n{prep_text}\n\n"


                        
                        webhook_url = current_webhook
                        if webhook_url:
                            print(f"DEBUG: 正在向 {webhook_url} 发送通知")
                            success, msg = send_dingtalk_notification(webhook_url, markdown_text, title=markdown_title)

                            
                            # 记录审计日志
                            audit = NotificationAudit(
                                task_id=task.id,
                                task_title=task.title,
                                robot_name=task.alert_robot,
                                webhook_url=webhook_url,
                                msg_type='markdown',
                                title=markdown_title,
                                content=markdown_text,
                                status='成功' if success else '失败',
                                error_msg=None if success else msg
                            )
                            db.session.add(audit)
                            
                            if success:
                                print(f"DEBUG: 通知发送成功")
                                # 处理周期性逻辑
                                if task.schedule_type == 'once':
                                    task.reminder_sent = True
                                else:
                                    # 使用 task.plan_time 作为基准，强制计算“下一个”周期
                                    next_run = calculate_next_run_time(task.plan_time, task.schedule_type, task.schedule_value, base_time=task.plan_time)
                                    if next_run:

                                        print(f"DEBUG: 周期任务[{task.title}]，更新计划时间从 {task.plan_time} 到 {next_run}")
                                        task.plan_time = next_run
                                        task.reminder_sent = False 
                                    else:
                                        task.reminder_sent = True
                                db.session.commit()
                            else:
                                print(f"DEBUG: 通知发送失败: {msg}")
                                # 即使失败也标记为已发送，防止阻塞
                                task.reminder_sent = True
                                db.session.commit()
                        else:
                            print(f"DEBUG: 任务[{task.title}] 未配置 Webhook，标记为已处理")
                            task.reminder_sent = True
                            db.session.commit()

                
            except Exception as e:
                print(f"ERROR: 后台提醒线程异常: {str(e)}")
                import traceback
                traceback.print_exc()
                db.session.rollback()

            
            time.sleep(60)

# 启动后台工作线程
start_background_worker()



# ==================== API接口 ====================






# 登录验证装饰器
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'code': -1, 'message': '请先登录'}), 401
        return f(*args, **kwargs)
    return decorated_function

# 管理员验证装饰器
def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'code': -1, 'message': '请先登录'}), 401
        if session.get('role') != 'admin':
            return jsonify({'code': -1, 'message': '权限不足，仅管理员可执行此操作'}), 403
        return f(*args, **kwargs)
    return decorated_function


@app.route('/')
def index():
    # 如果未登录，重定向到登录页
    if 'user_id' not in session:
        return send_from_directory('static', 'login.html')
    return send_from_directory('static', 'index.html')

@app.route('/login.html')
def login_page():
    return send_from_directory('static', 'login.html')

# 禁用静态文件缓存（开发环境）
@app.after_request
def add_header(response):
    if 'text/html' in response.content_type or 'application/javascript' in response.content_type or 'text/css' in response.content_type:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response


# 用户认证接口
@app.route('/api/login', methods=['POST'])
def login():
    """用户登录"""
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'code': -1, 'message': '用户名和密码不能为空'})
    
    user = User.query.filter_by(username=username).first()
    
    if not user or not user.check_password(password):
        return jsonify({'code': -1, 'message': '用户名或密码错误'})
    
    if not user.is_active:
        return jsonify({'code': -1, 'message': '账号已被禁用'})
    
    # 更新最后登录时间
    user.last_login = datetime.now()
    db.session.commit()
    
    # 设置session
    session.clear()
    session.permanent = True
    session['user_id'] = user.id
    session['username'] = user.username
    session['real_name'] = user.real_name
    session['role'] = user.role
    
    return jsonify({
        'code': 0,
        'message': '登录成功',
        'data': {
            'username': user.username,
            'real_name': user.real_name,
            'role': user.role
        }
    })

@app.route('/api/logout', methods=['POST'])
def logout():
    """用户登出"""
    session.clear()
    return jsonify({'code': 0, 'message': '登出成功'})

@app.route('/api/current-user', methods=['GET'])
def get_current_user():
    """获取当前登录用户"""
    if 'user_id' not in session:
        return jsonify({'code': -1, 'message': '未登录'})
    
    user = User.query.get(session['user_id'])
    if not user:
        session.clear()
        return jsonify({'code': -1, 'message': '用户不存在'})
    
    return jsonify({
        'code': 0,
        'data': {
            'id': user.id,
            'username': user.username,
            'real_name': user.real_name,
            'role': user.role,
            'department': user.department,
            'phone': user.phone,
            'email': user.email
        }
    })

# 用户管理接口
@app.route('/api/users', methods=['GET'])
@admin_required
def get_users():

    """获取用户列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    search = request.args.get('search', '')
    
    query = User.query
    if search:
        query = query.filter(
            db.or_(
                User.username.contains(search),
                User.real_name.contains(search),
                User.department.contains(search)
            )
        )
    
    pagination = query.order_by(User.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'code': 0,
        'data': {
            'items': [{
                'id': u.id,
                'username': u.username,
                'real_name': u.real_name,
                'role': u.role,
                'department': u.department,
                'phone': u.phone,
                'email': u.email,
                'is_active': u.is_active,
                'last_login': u.last_login.strftime('%Y-%m-%d %H:%M:%S') if u.last_login else None,
                'created_at': u.created_at.strftime('%Y-%m-%d %H:%M:%S')
            } for u in pagination.items],
            'total': pagination.total,
            'page': page,
            'per_page': per_page
        }
    })

@app.route('/api/users', methods=['POST'])
@admin_required
def create_user():
    """创建用户"""
    data = request.json

    
    if User.query.filter_by(username=data['username']).first():
        return jsonify({'code': -1, 'message': '用户名已存在'})
    
    user = User(
        username=data['username'],
        real_name=data['real_name'],
        role=data.get('role', 'user'),
        department=data.get('department'),
        phone=data.get('phone'),
        email=data.get('email'),
        is_active=data.get('is_active', True)
    )
    user.set_password(data['password'])
    
    db.session.add(user)
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '创建成功', 'data': {'id': user.id}})

@app.route('/api/users/<int:id>', methods=['PUT'])
@admin_required
def update_user(id):
    """更新用户"""
    user = User.query.get_or_404(id)
    data = request.json
    
    if 'username' in data and data['username'] != user.username:

        if User.query.filter_by(username=data['username']).first():
            return jsonify({'code': -1, 'message': '用户名已存在'})
        user.username = data['username']
    
    if 'real_name' in data:
        user.real_name = data['real_name']
    if 'role' in data:
        user.role = data['role']
    if 'department' in data:
        user.department = data['department']
    if 'phone' in data:
        user.phone = data['phone']
    if 'email' in data:
        user.email = data['email']
    if 'is_active' in data:
        user.is_active = data['is_active']
    if 'password' in data and data['password']:
        user.set_password(data['password'])
    
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '更新成功'})

@app.route('/api/users/<int:id>', methods=['DELETE'])
@admin_required
def delete_user(id):
    """删除用户"""
    # 不能删除自己
    if session.get('user_id') == id:
        return jsonify({'code': -1, 'message': '不能删除自己'})
    
    user = User.query.get_or_404(id)

    db.session.delete(user)
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '删除成功'})

# 业务系统相关接口
@app.route('/api/business-systems', methods=['GET'])
@login_required
def get_business_systems():
    """获取业务系统列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    search = request.args.get('search', '')
    status = request.args.get('status', '')
    
    query = BusinessSystem.query
    if search:
        # 同时匹配系统名称和IP地址
        query = query.join(SystemHost, isouter=True).filter(
            db.or_(
                BusinessSystem.system_name.contains(search),
                SystemHost.ip_address.contains(search)
            )
        )
    
    if status:
        query = query.filter(BusinessSystem.status == status)
    
    query = query.distinct()
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'code': 0,
        'data': {
            'items': [{
                'id': sys.id,
                'system_name': sys.system_name,
                'system_code': sys.system_code,
                'database': sys.database,
                'database_version': sys.database_version,
                'department': sys.department,
                'department_status': sys.department_status,
                'status': sys.status,
                'description': sys.description,
                'contact_person': sys.contact_person,
                'contact_phone': sys.contact_phone,
                'contact_email': sys.contact_email,
                'construction_unit': sys.construction_unit,
                'location': sys.location,
                'access_url': sys.access_url,

                'hosts': [{
                    'id': h.id,
                    'host_type': h.host_type,
                    'ip_address': h.ip_address,
                    'host_purpose': h.host_purpose,
                    'os_version': h.os_version,
                    'cpu_cores': h.cpu_cores,
                    'memory_gb': h.memory_gb,
                    'disk_gb': h.disk_gb,
                    'cpu_arch': h.cpu_arch
                } for h in sys.hosts],

                'middlewares': [{
                    'id': m.id,
                    'middleware_type': m.middleware_type,
                    'middleware_version': m.middleware_version,
                    'quantity': m.quantity
                } for m in sys.middlewares],
                'integrations': [{
                    'id': i.id,
                    'integration_type': i.integration_type,
                    'remote_system_name': i.remote_system_name,
                    'network_type': i.network_type
                } for i in sys.integrations],
                'created_at': sys.created_at.strftime('%Y-%m-%d %H:%M:%S') if sys.created_at else None

            } for sys in pagination.items],
            'total': pagination.total,
            'page': page,
            'per_page': per_page
        }
    })

@app.route('/api/business-systems/export', methods=['POST'])
@login_required
def export_business_systems():
    """导出业务系统到 Excel"""
    data = request.json
    ids = data.get('ids', [])
    
    query = BusinessSystem.query
    if ids:
        query = query.filter(BusinessSystem.id.in_(ids))
    
    systems = query.all()
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return jsonify({'code': -1, 'message': '未安装 openpyxl 库，无法导出'}), 500
        
    wb = Workbook()
    ws = wb.active
    ws.title = "业务系统清单"
    
    # 定义表头
    headers = [
        "系统名称", "系统代码", "状态", "建设单位", "管理部室", "部室状态", 
        "数据库类型", "数据库版本", "负责人", "联系电话", "联系邮箱", 
        "系统描述", "主机信息", "中间件信息", "创建时间"
    ]

    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        
    # 填充数据
    for row, sys in enumerate(systems, 2):
        # 格式化主机信息
        hosts_str = "\n".join([
            f"[{h.host_type}] {h.ip_address} ({h.os_version or '-'}, {h.cpu_arch or '-'}, {h.cpu_cores or '-'}核/{h.memory_gb or '-'}GB/{h.disk_gb or '-'}GB)"
            for h in sys.hosts
        ])
        
        # 格式化中间件信息
        mws_str = "\n".join([
            f"{m.middleware_type} {m.middleware_version or ''} (数量: {m.quantity})"
            for m in sys.middlewares
        ])
        
        ws.cell(row=row, column=1, value=sys.system_name)
        ws.cell(row=row, column=2, value=sys.system_code)
        ws.cell(row=row, column=3, value=sys.status)
        ws.cell(row=row, column=4, value=sys.construction_unit)
        ws.cell(row=row, column=5, value=sys.department)
        ws.cell(row=row, column=6, value=sys.department_status)
        ws.cell(row=row, column=7, value=sys.database)
        ws.cell(row=row, column=8, value=sys.database_version)
        ws.cell(row=row, column=9, value=sys.contact_person)
        ws.cell(row=row, column=10, value=sys.contact_phone)
        ws.cell(row=row, column=11, value=sys.contact_email)
        ws.cell(row=row, column=12, value=sys.description)
        ws.cell(row=row, column=13, value=hosts_str).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=14, value=mws_str).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=15, value=sys.created_at.strftime('%Y-%m-%d %H:%M:%S') if sys.created_at else "")
        
    # 设置列宽
    column_widths = [25, 15, 10, 20, 20, 15, 15, 15, 15, 15, 20, 30, 50, 40, 20]

    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # 保存到内存流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"business_systems_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/business-systems', methods=['POST'])

@login_required
def create_business_system():
    """创建业务系统"""
    data = request.json
    
    # 检查系统名称是否已存在
    if BusinessSystem.query.filter_by(system_name=data['system_name']).first():
        return jsonify({'code': -1, 'message': '系统名称已存在'}), 400
    
    system = BusinessSystem(
        system_name=data['system_name'],
        system_code=data.get('system_code'),
        database=data.get('database'),
        database_version=data.get('database_version'),
        department=data.get('department'),
        department_status=data.get('department_status'),
        status=data.get('status', '运行中'),
        description=data.get('description'),
        contact_person=data.get('contact_person'),
        contact_phone=data.get('contact_phone'),
        contact_email=data.get('contact_email'),
        construction_unit=data.get('construction_unit'),
        location=data.get('location'),
        access_url=data.get('access_url')
    )

    
    db.session.add(system)
    db.session.flush()
    
    # 添加主机信息
    if 'hosts' in data:
        for host_data in data['hosts']:
            host = SystemHost(
                system_id=system.id,
                host_type=host_data.get('host_type'),
                ip_address=host_data.get('ip_address'),
                host_purpose=host_data.get('host_purpose'),
                os_version=host_data.get('os_version'),
                cpu_cores=host_data.get('cpu_cores'),
                memory_gb=host_data.get('memory_gb'),
                disk_gb=host_data.get('disk_gb'),
                cpu_arch=host_data.get('cpu_arch')
            )
            db.session.add(host)
    
    # 添加中间件信息
    if 'middlewares' in data:
        for mw_data in data['middlewares']:
            middleware = SystemMiddleware(
                system_id=system.id,
                middleware_type=mw_data.get('middleware_type'),
                middleware_version=mw_data.get('middleware_version'),
                quantity=mw_data.get('quantity', 1)
            )
            db.session.add(middleware)

    # 添加集成信息（上/下游关联系统）
    if 'integrations' in data:
        for int_data in data['integrations']:
            integration = SystemIntegration(
                system_id=system.id,
                integration_type=int_data.get('integration_type'),
                remote_system_name=int_data.get('remote_system_name'),
                network_type=int_data.get('network_type')
            )
            db.session.add(integration)
    
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '创建成功', 'data': {'id': system.id}})



@app.route('/api/business-systems/<int:id>', methods=['PUT'])
@login_required
def update_business_system(id):
    """更新业务系统"""
    system = BusinessSystem.query.get_or_404(id)
    data = request.json
    
    # 检查系统名称是否与其他系统重复
    existing = BusinessSystem.query.filter_by(system_name=data['system_name']).first()
    if existing and existing.id != id:
        return jsonify({'code': -1, 'message': '系统名称已存在'}), 400
    
    system.system_name = data['system_name']
    system.system_code = data.get('system_code')
    system.database = data.get('database')
    system.database_version = data.get('database_version')
    system.department = data.get('department')
    system.department_status = data.get('department_status')
    system.status = data.get('status', '运行中')
    system.description = data.get('description')
    system.contact_person = data.get('contact_person')
    system.contact_phone = data.get('contact_phone')
    system.contact_email = data.get('contact_email')
    system.construction_unit = data.get('construction_unit')
    system.location = data.get('location')
    system.access_url = data.get('access_url')


    
    # 删除旧的主机信息
    SystemHost.query.filter_by(system_id=system.id).delete()
    
    # 添加新的主机信息
    if 'hosts' in data:
        for host_data in data['hosts']:
            host = SystemHost(
                system_id=system.id,
                host_type=host_data.get('host_type'),
                ip_address=host_data.get('ip_address'),
                host_purpose=host_data.get('host_purpose'),
                os_version=host_data.get('os_version'),
                cpu_cores=host_data.get('cpu_cores'),
                memory_gb=host_data.get('memory_gb'),
                disk_gb=host_data.get('disk_gb'),
                cpu_arch=host_data.get('cpu_arch')
            )
            db.session.add(host)

    
    # 删除旧的中间件信息
    SystemMiddleware.query.filter_by(system_id=system.id).delete()
    
    # 添加新的中间件信息
    if 'middlewares' in data:
        for mw_data in data['middlewares']:
            middleware = SystemMiddleware(
                system_id=system.id,
                middleware_type=mw_data.get('middleware_type'),
                middleware_version=mw_data.get('middleware_version'),
                quantity=mw_data.get('quantity', 1)
            )
            db.session.add(middleware)
    
    # 删除旧的集成信息
    SystemIntegration.query.filter_by(system_id=system.id).delete()
    
    # 添加新的集成信息
    if 'integrations' in data:
        for int_data in data['integrations']:
            integration = SystemIntegration(
                system_id=system.id,
                integration_type=int_data.get('integration_type'),
                remote_system_name=int_data.get('remote_system_name'),
                network_type=int_data.get('network_type')
            )
            db.session.add(integration)
    
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '更新成功'})


@app.route('/api/business-systems/<int:id>', methods=['DELETE'])
@login_required
def delete_business_system(id):
    """删除业务系统"""
    system = BusinessSystem.query.get_or_404(id)
    
    # 检查是否有关联的事件
    if system.events:
        return jsonify({'code': -1, 'message': '该系统存在关联事件,无法删除'}), 400
    
    db.session.delete(system)
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '删除成功'})

# 事件相关接口
@app.route('/api/events', methods=['GET'])
@login_required
def get_events():
    """获取事件列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    system_name = request.args.get('system_name', '')
    title = request.args.get('title', '')
    status = request.args.get('status', '')
    progress_status = request.args.get('progress_status', '')
    event_type = request.args.get('event_type', '')
    severity = request.args.get('severity', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    query = Event.query
    
    if system_name:
        query = query.filter(Event.system_name.contains(system_name))
    if title:
        query = query.filter(Event.title.contains(title))
    if status:
        query = query.filter_by(status=status)
    if progress_status:
        query = query.filter_by(progress_status=progress_status)
    if event_type:
        query = query.filter_by(event_type=event_type)

    if severity:
        query = query.filter_by(severity=severity)
    if start_date:
        query = query.filter(Event.occurred_at >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(Event.occurred_at <= datetime.fromisoformat(end_date))
    
    query = query.order_by(Event.occurred_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    
    return jsonify({
        'code': 0,
        'data': {
            'items': [{
                'id': event.id,
                'event_no': event.event_no,
                'system_name': event.system_name,
                'event_type': event.event_type,
                'event_category': event.event_category,
                'severity': event.severity,
                'status': event.status,
                'title': event.title,
                'description': event.description,
                'occurred_at': event.occurred_at.strftime('%Y-%m-%d %H:%M:%S') if event.occurred_at else None,
                'reported_by': event.reported_by,
                'assigned_to': event.assigned_to,
                'progress_status': event.progress_status,
                'created_at': event.created_at.strftime('%Y-%m-%d %H:%M:%S') if event.created_at else None
            } for event in pagination.items],

            'total': pagination.total,
            'page': page,
            'per_page': per_page
        }
    })

@app.route('/api/events/<int:id>', methods=['GET'])
@login_required
def get_event_detail(id):
    """获取事件详情"""
    event = Event.query.get_or_404(id)
    
    return jsonify({
        'code': 0,
        'data': {
            'id': event.id,
            'event_no': event.event_no,
            'system_id': event.system_id,
            'system_name': event.system_name,
            'event_type': event.event_type,
            'event_category': event.event_category,
            'severity': event.severity,
            'status': event.status,
            'title': event.title,
            'description': event.description,
            'occurred_at': event.occurred_at.strftime('%Y-%m-%d %H:%M:%S') if event.occurred_at else None,
            'reported_by': event.reported_by,
            'assigned_to': event.assigned_to,
            'progress_status': event.progress_status,
            'resolution': event.resolution,
            'root_cause': event.root_cause,

            'resolved_at': event.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if event.resolved_at else None,
            'closed_at': event.closed_at.strftime('%Y-%m-%d %H:%M:%S') if event.closed_at else None,
            'processes': [{
                'id': p.id,
                'step_no': p.step_no,
                'action': p.action,
                'result': p.result,
                'operator': p.operator,
                'operated_at': p.operated_at.strftime('%Y-%m-%d %H:%M:%S') if p.operated_at else None,
                'remarks': p.remarks
            } for p in event.processes],
            'attachments': [{
                'id': a.id,
                'file_name': a.file_name,
                'file_type': a.file_type,
                'file_size': a.file_size,
                'uploaded_by': a.uploaded_by,
                'uploaded_at': a.uploaded_at.strftime('%Y-%m-%d %H:%M:%S')
            } for a in event.attachments]
        }
    })

@app.route('/api/events', methods=['POST'])
@login_required
def create_event():
    """创建事件"""
    data = request.json
    
    # 生成事件编号
    today = datetime.now().strftime('%Y%m%d')
    count = Event.query.filter(Event.event_no.like(f'INC-{today}%')).count()
    event_no = f'INC-{today}-{count + 1:04d}'
    
    # 获取系统信息
    system = BusinessSystem.query.get(data['system_id'])
    if not system:
        return jsonify({'code': -1, 'message': '业务系统不存在'}), 400
    
    event = Event(
        event_no=event_no,
        system_id=data['system_id'],
        system_name=system.system_name,
        event_type=data['event_type'],
        severity=data.get('severity'),

        status=data.get('status', '处理中'),
        title=data['title'],
        description=data.get('description'),
        occurred_at=datetime.fromisoformat(data['occurred_at']),
        reported_by=data.get('reported_by'),
        assigned_to=data.get('assigned_to'),
        progress_status=data.get('progress_status', '未解决'),
        resolution=data.get('resolution'),

        root_cause=data.get('root_cause')
    )
    
    db.session.add(event)
    db.session.flush()  # 获取event.id
    
    # 添加处置流程
    if 'processes' in data:
        for idx, process_data in enumerate(data['processes']):
            process = EventProcess(
                event_id=event.id,
                step_no=idx + 1,
                action=process_data['action'],
                result=process_data.get('result'),
                operator=process_data.get('operator'),
                operated_at=datetime.fromisoformat(process_data.get('operated_at')) if process_data.get('operated_at') else datetime.now(),
                remarks=process_data.get('remarks')

            )
            db.session.add(process)
    
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '创建成功', 'data': {'id': event.id, 'event_no': event_no}})

@app.route('/api/events/<int:id>', methods=['PUT'])
@login_required
def update_event(id):
    """更新事件"""
    event = Event.query.get_or_404(id)
    data = request.json
    
    # 获取系统信息
    if 'system_id' in data:
        system = BusinessSystem.query.get(data['system_id'])
        if not system:
            return jsonify({'code': -1, 'message': '业务系统不存在'}), 400
        event.system_id = data['system_id']
        event.system_name = system.system_name
    
    if 'event_type' in data:
        event.event_type = data['event_type']
    if 'severity' in data:

        event.severity = data['severity']
    if 'status' in data:
        event.status = data['status']
        if data['status'] == '已解决' and not event.resolved_at:
            event.resolved_at = datetime.now()
        elif data['status'] == '已关闭' and not event.closed_at:
            event.closed_at = datetime.now()
    if 'title' in data:
        event.title = data['title']
    if 'description' in data:
        event.description = data['description']
    if 'occurred_at' in data:
        event.occurred_at = datetime.fromisoformat(data['occurred_at'])
    if 'reported_by' in data:
        event.reported_by = data['reported_by']
    if 'assigned_to' in data:
        event.assigned_to = data['assigned_to']
    if 'progress_status' in data:
        event.progress_status = data['progress_status']
    if 'resolution' in data:
        event.resolution = data['resolution']

    if 'root_cause' in data:
        event.root_cause = data['root_cause']
    
    # 更新处置流程
    if 'processes' in data:
        # 删除旧的流程
        EventProcess.query.filter_by(event_id=event.id).delete()
        
        # 添加新的流程
        for idx, process_data in enumerate(data['processes']):
            process = EventProcess(
                event_id=event.id,
                step_no=idx + 1,
                action=process_data['action'],
                result=process_data.get('result'),
                operator=process_data.get('operator'),
                operated_at=datetime.fromisoformat(process_data.get('operated_at')) if process_data.get('operated_at') else datetime.now(),
                remarks=process_data.get('remarks')

            )
            db.session.add(process)
    
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '更新成功'})

@app.route('/api/events/<int:id>', methods=['DELETE'])
@login_required
def delete_event(id):
    """删除事件"""
    event = Event.query.get_or_404(id)
    
    # 删除关联的附件文件
    for attachment in event.attachments:
        try:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], attachment.file_path)
            if os.path.exists(file_path):
                os.remove(file_path)
        except:
            pass
    
    db.session.delete(event)
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '删除成功'})

# 附件上传接口
@app.route('/api/events/<int:event_id>/attachments', methods=['POST'])
@login_required
def upload_attachment(event_id):
    """上传附件"""
    event = Event.query.get_or_404(event_id)
    
    if 'file' not in request.files:
        return jsonify({'code': -1, 'message': '没有文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': -1, 'message': '没有选择文件'}), 400
    
    filename = secure_filename(file.filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{timestamp}_{filename}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)
    
    attachment = EventAttachment(
        event_id=event_id,
        file_name=file.filename,
        file_path=filename,
        file_type=file.content_type,
        file_size=os.path.getsize(file_path),
        uploaded_by=request.form.get('uploaded_by', '未知')
    )
    
    db.session.add(attachment)
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '上传成功', 'data': {'id': attachment.id}})

@app.route('/api/attachments/<int:id>', methods=['DELETE'])
@login_required
def delete_attachment(id):
    """删除附件"""
    attachment = EventAttachment.query.get_or_404(id)
    
    # 删除文件
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], attachment.file_path)
        if os.path.exists(file_path):
            os.remove(file_path)
    except:
        pass
    
    db.session.delete(attachment)
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '删除成功'})

@app.route('/api/attachments/<int:id>/download', methods=['GET'])
@login_required
def download_attachment(id):
    """下载附件"""
    attachment = EventAttachment.query.get_or_404(id)
    return send_from_directory(app.config['UPLOAD_FOLDER'], attachment.file_path, as_attachment=True, download_name=attachment.file_name)

# 统计相关接口
@app.route('/api/dashboard/overview', methods=['GET'])
@login_required
def get_dashboard_overview():
    """获取概览统计数据"""
    # 业务系统数量
    total_systems = BusinessSystem.query.count()
    
    # 事件数量统计
    total_events = Event.query.count()
    pending_events = Event.query.filter_by(status='待处理').count()
    processing_events = Event.query.filter_by(status='处理中').count()
    resolved_events = Event.query.filter_by(status='已解决').count()
    closed_events = Event.query.filter_by(status='已关闭').count()
    
    # 按状态分类统计
    status_stats = [
        {'status': '待处理', 'count': pending_events},
        {'status': '处理中', 'count': processing_events},
        {'status': '已解决', 'count': resolved_events},
        {'status': '已关闭', 'count': closed_events}
    ]
    
    # 按事件类型统计
    type_stats = db.session.query(
        Event.event_type, 
        db.func.count(Event.id)
    ).group_by(Event.event_type).all()
    
    # 按严重程度统计
    severity_stats = db.session.query(
        Event.severity, 
        db.func.count(Event.id)
    ).group_by(Event.severity).all()
    
    # 最近事件列表
    recent_events = Event.query.order_by(Event.created_at.desc()).limit(5).all()
    
    # 平均响应时间(小时)
    resolved = Event.query.filter(Event.resolved_at.isnot(None)).all()
    avg_response_time = 0
    if resolved:
        total_time = sum([(e.resolved_at - e.occurred_at).total_seconds() for e in resolved])
        avg_response_time = round(total_time / len(resolved) / 3600, 1)
    
    # 计划任务统计
    total_tasks = PlanTask.query.count()
    completed_tasks = PlanTask.query.filter_by(status='已完成').count()
    pending_tasks = total_tasks - completed_tasks
    
    return jsonify({
        'code': 0,
        'data': {
            'total_systems': total_systems,
            'total_events': total_events,
            'pending_events': pending_events,
            'processing_events': processing_events,
            'avg_response_time': avg_response_time,
            'status_stats': status_stats,
            'type_stats': [{'type': t[0], 'count': t[1]} for t in type_stats],
            'severity_stats': [{'severity': s[0], 'count': s[1]} for s in severity_stats],
            'recent_events': [{
                'id': e.id,
                'event_no': e.event_no,
                'system_name': e.system_name,
                'event_type': e.event_type,
                'severity': e.severity,
                'status': e.status,
                'occurred_at': e.occurred_at.strftime('%Y-%m-%d %H:%M:%S')
            } for e in recent_events],
            'plan_task_stats': {
                'completed': completed_tasks,
                'pending': pending_tasks
            }
        }
    })

# Dashboard趋势数据API
@app.route('/api/dashboard/trend', methods=['GET'])
@login_required
def get_dashboard_trend():
    """获取事件趋势数据"""
    from datetime import timedelta
    
    period = request.args.get('period', 'week')  # today, week, month
    now = datetime.now()
    
    if period == 'today':
        # 今日按小时统计
        start_time = now.replace(hour=0, minute=0, second=0, microsecond=0)
        labels = [f"{i}时" for i in range(24)]
        data = []
        for i in range(24):
            hour_start = start_time + timedelta(hours=i)
            hour_end = hour_start + timedelta(hours=1)
            count = Event.query.filter(
                Event.occurred_at >= hour_start,
                Event.occurred_at < hour_end
            ).count()
            data.append(count)
    
    elif period == 'week':
        # 本周按天统计
        start_time = now - timedelta(days=6)
        start_time = start_time.replace(hour=0, minute=0, second=0, microsecond=0)
        labels = []
        data = []
        for i in range(7):
            day_start = start_time + timedelta(days=i)
            day_end = day_start + timedelta(days=1)
            count = Event.query.filter(
                Event.occurred_at >= day_start,
                Event.occurred_at < day_end
            ).count()
            data.append(count)
            labels.append(day_start.strftime('%m-%d'))
    
    else:  # month
        # 本月按天统计
        start_time = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        labels = []
        data = []
        current_day = start_time
        while current_day.month == now.month:
            day_end = current_day + timedelta(days=1)
            count = Event.query.filter(
                Event.occurred_at >= current_day,
                Event.occurred_at < day_end
            ).count()
            data.append(count)
            labels.append(current_day.strftime('%d日'))
            current_day = day_end
    
    return jsonify({
        'code': 0,
        'data': {
            'labels': labels,
            'data': data
        }
    })

# 用户登录统计API
@app.route('/api/dashboard/login-stats', methods=['GET'])
@login_required
def get_login_stats():
    """获取用户登录统计（最近7天）"""
    from datetime import timedelta
    
    # 获取最近7天有登录的用户
    seven_days_ago = datetime.now() - timedelta(days=7)
    users = User.query.filter(
        User.last_login.isnot(None),
        User.last_login >= seven_days_ago
    ).order_by(User.last_login.desc()).limit(10).all()
    
    stats = []
    for user in users:
        # 计算登录次数（这里简化处理，实际应该有登录日志表）
        login_count = 1  # 简化处理
        stats.append({
            'username': user.username,
            'real_name': user.real_name,
            'role': user.role,
            'login_count': login_count,
            'last_login': user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else None
        })
    
    return jsonify({
        'code': 0,
        'data': stats
    })

# 计划任务管理接口
@app.route('/api/plan-tasks', methods=['GET'])
@login_required
def list_plan_tasks():
    """获取计划任务列表"""
    view = request.args.get('view', 'all')  # mine/created/all
    status = request.args.get('status', '')
    task_type = request.args.get('task_type', '')
    keyword = request.args.get('keyword', '')
    
    query = PlanTask.query
    current_user = session.get('username')
    
    if view == 'mine' and current_user:
        query = query.filter(
            db.or_(
                PlanTask.owner == current_user,
                PlanTask.responsible.contains(current_user)
            )
        )
    elif view == 'created' and current_user:
        query = query.filter(PlanTask.created_by == current_user)
    
    if status:
        query = query.filter(PlanTask.status == status)
    if task_type:
        query = query.filter(PlanTask.task_type == task_type)
    if keyword:
        query = query.filter(PlanTask.title.contains(keyword))
    
    tasks = query.order_by(PlanTask.plan_time.asc()).all()
    
    return jsonify({
        'code': 0,
        'data': {
            'items': [serialize_plan_task(task, simple=True) for task in tasks],
            'total': len(tasks)
        }
    })

@app.route('/api/plan-tasks/<int:task_id>', methods=['GET'])
@login_required
def get_plan_task(task_id):
    task = PlanTask.query.get_or_404(task_id)
    return jsonify({'code': 0, 'data': serialize_plan_task(task)})

@app.route('/api/notification-audits', methods=['GET'])
@login_required
def get_notification_audits():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 15, type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    task_title = request.args.get('task_title')
    
    query = NotificationAudit.query
    
    if start_date:
        query = query.filter(NotificationAudit.sent_at >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        # 包含结束当天
        query = query.filter(NotificationAudit.sent_at <= datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
    if task_title:
        query = query.filter(NotificationAudit.task_title.like(f"%{task_title}%"))
        
    pagination = query.order_by(NotificationAudit.sent_at.desc()).paginate(page=page, per_page=per_page)
    
    audits = []
    for audit in pagination.items:
        audits.append({
            'id': audit.id,
            'task_id': audit.task_id,
            'task_title': audit.task_title,
            'robot_name': audit.robot_name,
            'webhook_url': audit.webhook_url,
            'msg_type': audit.msg_type,
            'title': audit.title,
            'content': audit.content,
            'status': audit.status,
            'error_msg': audit.error_msg,
            'sent_at': audit.sent_at.strftime('%Y-%m-%d %H:%M:%S')
        })
        
    return jsonify({
        'code': 0,
        'data': audits,
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': pagination.page
    })

@app.route('/api/notification-audits/bulk-delete', methods=['POST'])
@admin_required
def bulk_delete_notification_audits():
    """批量删除通知审计记录"""
    data = request.json
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'code': -1, 'message': '请选择要删除的记录'})
    
    try:
        NotificationAudit.query.filter(NotificationAudit.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'code': 0, 'message': f'成功删除 {len(ids)} 条记录'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': -1, 'message': f'删除失败: {str(e)}'})

@app.route('/api/plan-tasks/test-notification', methods=['POST'])

@login_required
def test_plan_task_notification():
    data = request.json
    webhook_url = data.get('webhook_url')
    template = data.get('reminder_message') or '【计划任务提醒】任务：{title}，计划时间：{plan_time}，负责人：{owner}。'
    
    # 模拟或获取变量
    title = data.get('title', '测试任务')
    plan_time = data.get('plan_time', datetime.now().strftime('%Y-%m-%d %H:%M'))
    owner = data.get('owner', '测试负责人')
    responsible = data.get('responsible', [])
    if isinstance(responsible, list):
        responsible = '、'.join(responsible)
    
    preps = data.get('preparations', [])
    completed_count = len([p for p in preps if p.get('status') == '已完成'])
    # 使用符号更直观地展示状态
    prep_text_list = []
    for p in preps:
        status_icon = "✅" if p.get('status') == '已完成' else "⬜"
        prep_text_list.append(f"{status_icon} {p.get('description')}")
    
    prep_text = '\n\n'.join(prep_text_list) if prep_text_list else '无'
    prep_progress = f"{completed_count}/{len(preps)}"
    
    # 处理换行符：将字面量 \n 替换为实际换行符，并统一换行格式
    template = template.replace('\\n', '\n').replace('\r\n', '\n')
    
    # 替换变量
    message = template
    replacements = {
        '{title}': title,
        '{plan_time}': plan_time,
        '{owner}': owner,
        '{responsible}': responsible,
        '{preparations}': prep_text,
        '{prep_progress}': prep_progress
    }
    
    for key, val in replacements.items():
        message = message.replace(key, str(val))
    
    # 构建美化的 Markdown 消息
    markdown_title = f"⏰ 计划任务提醒: {title}"
    # 确保引用块中的每一行都带上 > 符号，且 Markdown 换行建议使用双换行或末尾双空格
    safe_message = message.replace('\n', '\n\n> ')
    markdown_text = f"### ⏰ 计划任务提醒\n\n" \
                    f"**任务名称**: <font color='#1d4ed8'>{title}</font>\n\n" \
                    f"--- \n\n" \
                    f"📅 **计划时间**: {plan_time}\n\n" \
                    f"👤 **主负责人**: {owner}\n\n" \
                    f"👥 **责任人**: {responsible}\n\n" \
                    f"📊 **当前进度**: `{prep_progress}`\n\n" \
                    f"📝 **准备事项**:\n\n{prep_text}\n\n"



    success, msg = send_dingtalk_notification(webhook_url, markdown_text, title=markdown_title)
    
    # 记录测试通知审计日志
    audit = NotificationAudit(
        task_id=data.get('id'), # 如果是新建任务可能没有ID
        task_title=title,
        robot_name=data.get('alert_robot', '测试机器人'),
        webhook_url=webhook_url,
        msg_type='markdown',
        title=markdown_title,
        content=markdown_text,
        status='成功' if success else '失败',
        error_msg=None if success else msg
    )
    db.session.add(audit)
    db.session.commit()
    
    if success:

        return jsonify({'code': 0, 'message': msg})
    else:
        return jsonify({'code': -1, 'message': msg})

@app.route('/api/plan-tasks', methods=['POST'])
@login_required
def create_plan_task():
    data = request.json
    try:
        plan_time = datetime.fromisoformat(data['plan_time'])
    except (KeyError, ValueError):
        return jsonify({'code': -1, 'message': '计划时间格式不正确'}), 400
    
    # 打印创建日志，检查 webhook 是否传入
    print(f"DEBUG: 创建任务 - Robot: {data.get('alert_robot')}, Webhook: {data.get('webhook_url')}")
    
    if data.get('schedule_type') in ['daily', 'weekly', 'monthly', 'cron']:
        plan_time = calculate_next_run_time(plan_time, data.get('schedule_type'), data.get('schedule_value'))

        # 回退一个周期，因为 calculate_next_run_time 总是寻找“下一个”
        # 但如果是新创建，我们可能希望如果是“今天的10点”且现在是9点，就用今天。
        # 修正：重新逻辑处理，如果 plan_time 本身就是未来的且符合要求，则不应强制跳到下一个。
        # 我已经把 calculate_next_run_time 改为只要 next_time <= now 就循环。
        # 这意味着如果传入的是“今天10点”且现在是9点，它不会进入循环，返回的就是今天10点。这符合预期。
    
    task = PlanTask(


        title=data.get('title'),
        task_type=data.get('task_type', '其他'),
        schedule_type=data.get('schedule_type', 'once'),
        schedule_value=data.get('schedule_value'),
        plan_time=plan_time,
        reminder_minutes=data.get('reminder_minutes', 1440),
        reminder_enabled=data.get('reminder_enabled', True),
        reminder_sent=False,
        alert_robot=data.get('alert_robot', '默认钉钉机器人'),

        webhook_url=data.get('webhook_url'),
        reminder_message=data.get('reminder_message') or '【计划任务提醒】任务：{title}，计划时间：{plan_time}，负责人：{owner}。请提前准备：{preparations}',
        status=data.get('status', '待执行'),
        responsible=','.join(data.get('responsible', [])),
        owner=data.get('owner'),
        description=data.get('description'),

        created_by=session.get('username')
    )
    
    db.session.add(task)
    db.session.flush()
    
    for index, item in enumerate(data.get('preparations', []), start=1):
        prep = PlanTaskPreparation(
            task_id=task.id,
            description=item.get('description', ''),
            status=item.get('status', '未开始'),
            estimated_minutes=item.get('estimated_minutes'),
            order_no=index
        )
        db.session.add(prep)
    
    db.session.commit()
    return jsonify({'code': 0, 'message': '创建成功', 'data': {'id': task.id}})

@app.route('/api/plan-tasks/<int:task_id>', methods=['PUT'])
@login_required
def update_plan_task(task_id):
    task = PlanTask.query.get_or_404(task_id)
    data = request.json
    
    if 'title' in data:
        task.title = data['title']
    if 'task_type' in data:
        task.task_type = data['task_type']
    if 'schedule_type' in data:
        task.schedule_type = data['schedule_type']
    if 'schedule_value' in data:
        task.schedule_value = data['schedule_value']
    if 'plan_time' in data:
        try:
            new_plan_time = datetime.fromisoformat(data['plan_time'])
            if task.schedule_type in ['daily', 'weekly', 'monthly', 'cron']:
                new_plan_time = calculate_next_run_time(new_plan_time, task.schedule_type, task.schedule_value)

            
            if new_plan_time != task.plan_time:

                task.plan_time = new_plan_time
                task.reminder_sent = False # 重置提醒状态
        except ValueError:
            return jsonify({'code': -1, 'message': '计划时间格式不正确'}), 400
    if 'reminder_minutes' in data:
        if task.reminder_minutes != data['reminder_minutes']:
            task.reminder_minutes = data['reminder_minutes']
            task.reminder_sent = False

    if 'reminder_enabled' in data:
        task.reminder_enabled = data['reminder_enabled']
    
    # 强制更新 Webhook URL，确保即使前端传空（如果配置里有）也能存入
    if 'alert_robot' in data:
        task.alert_robot = data['alert_robot']
    if 'webhook_url' in data:
        task.webhook_url = data['webhook_url']

    if 'reminder_message' in data:
        task.reminder_message = data['reminder_message']

    if 'responsible' in data:
        task.responsible = ','.join(data['responsible'])
    if 'owner' in data:
        task.owner = data['owner']
    if 'description' in data:
        task.description = data['description']
    if 'status' in data:
        task.status = data['status']
    if 'result_status' in data:
        task.result_status = data['result_status']
    if 'result_notes' in data:
        task.result_notes = data['result_notes']
    if 'actual_start' in data:
        task.actual_start = datetime.fromisoformat(data['actual_start']) if data['actual_start'] else None
    if 'actual_finish' in data:
        task.actual_finish = datetime.fromisoformat(data['actual_finish']) if data['actual_finish'] else None
    
    if 'preparations' in data:
        # 清空旧数据
        PlanTaskPreparation.query.filter_by(task_id=task.id).delete()
        db.session.flush()
        for index, item in enumerate(data['preparations'], start=1):
            prep = PlanTaskPreparation(
                task_id=task.id,
                description=item.get('description', ''),
                status=item.get('status', '未开始'),
                estimated_minutes=item.get('estimated_minutes'),
                order_no=index
            )
            db.session.add(prep)
    
    db.session.commit()
    return jsonify({'code': 0, 'message': '更新成功'})

@app.route('/api/plan-tasks/<int:task_id>/status', methods=['POST'])
@login_required
def update_plan_task_status(task_id):
    task = PlanTask.query.get_or_404(task_id)
    data = request.json or {}
    action = data.get('action')
    now = datetime.now()
    
    if action == 'start':
        task.status = '进行中'
        task.actual_start = now
    elif action == 'complete':
        task.status = '已完成'
        task.actual_finish = now
        task.result_status = data.get('result_status', '成功')
        task.result_notes = data.get('result_notes')
    elif action == 'cancel':
        task.status = '已取消'
        task.result_notes = data.get('result_notes')
    else:
        return jsonify({'code': -1, 'message': '未知的操作'}), 400
    
    db.session.commit()
    return jsonify({'code': 0, 'message': '状态更新成功'})


@app.route('/api/plan-tasks/<int:task_id>', methods=['DELETE'])
@login_required
def delete_plan_task(task_id):
    """删除计划任务"""
    task = PlanTask.query.get_or_404(task_id)
    db.session.delete(task)
    db.session.commit()
    return jsonify({'code': 0, 'message': '删除成功'})


def serialize_plan_task(task: PlanTask, simple=False):
    base = {
        'id': task.id,
        'title': task.title,
        'task_type': task.task_type,
        'schedule_type': task.schedule_type,
        'schedule_value': task.schedule_value,
        'plan_time': task.plan_time.strftime('%Y-%m-%d %H:%M'),
        'reminder_minutes': task.reminder_minutes,
        'reminder_enabled': task.reminder_enabled,
        'reminder_sent': task.reminder_sent,
        'alert_robot': task.alert_robot,

        'webhook_url': task.webhook_url,
        'reminder_message': task.reminder_message,
        'status': task.status,
        'responsible': task.responsible.split(',') if task.responsible else [],
        'owner': task.owner,
        'description': task.description,

        'created_by': task.created_by,
        'actual_start': task.actual_start.strftime('%Y-%m-%d %H:%M') if task.actual_start else None,
        'actual_finish': task.actual_finish.strftime('%Y-%m-%d %H:%M') if task.actual_finish else None,
        'result_status': task.result_status,
        'result_notes': task.result_notes,
        'created_at': task.created_at.strftime('%Y-%m-%d %H:%M')
    }
    
    if simple:
        return base
    
    base['preparations'] = [{
        'id': prep.id,
        'description': prep.description,
        'status': prep.status,
        'estimated_minutes': prep.estimated_minutes,
        'order_no': prep.order_no
    } for prep in sorted(task.preparations, key=lambda x: x.order_no)]
    return base

# 系统配置相关接口
@app.route('/api/configs', methods=['GET'])
@login_required
def get_configs():
    """获取系统配置"""
    config_type = request.args.get('config_type', '')
    
    query = SystemConfig.query
    if config_type:
        # config_type参数实际对应config_key字段
        query = query.filter_by(config_key=config_type)
    
    configs = query.all()
    
    return jsonify({
        'code': 0,
        'data': [{
            'id': c.id,
            'config_key': c.config_key,
            'config_value': c.config_value,
            'config_type': c.config_type,
            'description': c.description,
            'updated_at': c.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        } for c in configs]
    })

@app.route('/api/configs', methods=['POST'])
@admin_required
def create_config():

    """创建配置"""
    data = request.json
    
    if SystemConfig.query.filter_by(config_key=data['config_key']).first():
        return jsonify({'code': -1, 'message': '配置键已存在'}), 400
    
    config = SystemConfig(
        config_key=data['config_key'],
        config_value=data['config_value'],
        config_type=data.get('config_type'),
        description=data.get('description')
    )
    
    db.session.add(config)
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '创建成功', 'data': {'id': config.id}})

@app.route('/api/configs/<int:id>', methods=['PUT'])
@admin_required
def update_config(id):
    """更新配置"""
    config = SystemConfig.query.get_or_404(id)
    data = request.json
    
    config.config_value = data['config_value']
    if 'description' in data:
        config.description = data['description']
    
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '更新成功'})

@app.route('/api/configs/<int:id>', methods=['DELETE'])
@admin_required
def delete_config(id):

    """删除配置"""
    config = SystemConfig.query.get_or_404(id)
    
    db.session.delete(config)
    db.session.commit()
    
    return jsonify({'code': 0, 'message': '删除成功'})

# ==================== 初始化数据 ====================

def init_database():
    """初始化数据库和示例数据"""
    with app.app_context():
        # 创建表
        db.create_all()
        
        # 检查是否已有数据
        if BusinessSystem.query.first():
            return
        
        # 插入示例业务系统
        systems = [
            BusinessSystem(
                system_name='核心业务系统',
                system_code='CORE-SYS',
                database='Oracle',
                database_version='12c',
                department='运营中心',
                department_status='运行中',
                status='运行中',
                description='公司核心业务系统,承载主要业务流程',
                contact_person='张三',
                contact_phone='13800138000',
                contact_email='zhangsan@example.com'
            ),
            BusinessSystem(
                system_name='用户管理系统',
                system_code='USER-MGR',
                database='MySQL',
                database_version='8.0',
                department='技术中心',
                department_status='运行中',
                status='运行中',
                contact_person='李四',
                contact_phone='13900139000',
                contact_email='lisi@example.com'
            )
        ]
        
        for sys in systems:
            db.session.add(sys)
        
        db.session.flush()
        
        # 添加主机信息
        hosts = [
            SystemHost(system_id=systems[0].id, host_type='物理服务器', ip_address='192.168.1.100', host_purpose='应用服务器'),
            SystemHost(system_id=systems[0].id, host_type='物理服务器', ip_address='192.168.1.101', host_purpose='数据库服务器'),
            SystemHost(system_id=systems[1].id, host_type='云服务器', ip_address='192.168.1.102', host_purpose='Web服务器'),
        ]
        
        for host in hosts:
            db.session.add(host)
        
        # 添加中间件信息
        middlewares = [
            SystemMiddleware(system_id=systems[0].id, middleware_type='WebLogic', middleware_version='12c', quantity=2),
            SystemMiddleware(system_id=systems[1].id, middleware_type='Tomcat', middleware_version='9', quantity=3),
        ]
        
        for mw in middlewares:
            db.session.add(mw)
        
        db.session.commit()
        
        # 插入示例事件
        events_data = [
            {
                'system': systems[0],
                'event_type': '系统故障',
                'event_category': '数据库故障',
                'severity': '紧急',
                'status': '已解决',
                'title': 'Oracle数据库连接超时',
                'description': '核心业务系统无法连接Oracle数据库,导致业务中断',
                'occurred_at': datetime(2023, 5, 12, 14, 30),
                'reported_by': '张三',
                'assigned_to': '运维团队',
                'resolution': '重启数据库监听服务,恢复连接',
                'root_cause': '数据库监听服务异常停止',
                'resolved_at': datetime(2023, 5, 12, 15, 0)
            },
            {
                'system': systems[1],
                'event_type': '性能问题',
                'event_category': '响应缓慢',
                'severity': '一般',
                'status': '处理中',
                'title': '用户登录响应缓慢',
                'description': '用户反馈登录时响应时间超过10秒',
                'occurred_at': datetime(2023, 6, 15, 9, 15),
                'reported_by': '李四',
                'assigned_to': '开发团队'
            },
            {
                'system': systems[0],
                'event_type': '安全事件',
                'event_category': '异常访问',
                'severity': '严重',
                'status': '待处理',
                'title': '检测到异常数据访问',
                'description': '监控系统发现大量异常数据查询请求',
                'occurred_at': datetime(2023, 6, 19, 11, 0),
                'reported_by': '安全系统',
                'assigned_to': '安全团队'
            }
        ]
        
        for event_data in events_data:
            system = event_data.pop('system')
            resolved_at = event_data.pop('resolved_at', None)
            
            today = datetime.now().strftime('%Y%m%d')
            count = Event.query.count()
            event_no = f'INC-{today}-{count + 1:04d}'
            
            event = Event(
                event_no=event_no,
                system_id=system.id,
                system_name=system.system_name,
                **event_data
            )
            
            if resolved_at:
                event.resolved_at = resolved_at
            
            db.session.add(event)
            db.session.flush()
            
            # 添加处置流程
            if event.status in ['处理中', '已解决']:
                process1 = EventProcess(
                    event_id=event.id,
                    step_no=1,
                    action='接收事件并初步分析',
                    result='确认问题原因',
                    operator='运维人员',
                    remarks='开始处理'
                )
                db.session.add(process1)
                
                if event.status == '已解决':
                    process2 = EventProcess(
                        event_id=event.id,
                        step_no=2,
                        action='执行解决方案',
                        result='问题已解决',
                        operator='运维人员',
                        remarks='完成处理'
                    )
                    db.session.add(process2)
        
        # 插入系统配置
        configs = [
            SystemConfig(
                config_key='event_types',
                config_value='系统故障,性能问题,安全事件,变更请求,咨询问题',
                config_type='事件类型',
                description='事件类型选项'
            ),
            SystemConfig(
                config_key='severity_levels',
                config_value='紧急,严重,一般,较低',
                config_type='严重程度',
                description='严重程度选项'
            ),
            SystemConfig(
                config_key='event_status',
                config_value='待处理,处理中,已解决,已关闭',
                config_type='事件状态',
                description='事件状态选项'
            ),
            SystemConfig(
                config_key='departments',
                config_value='运营中心,技术中心,数据中心,运维中心,安全中心',
                config_type='管理部室',
                description='管理部室选项'
            ),
            SystemConfig(
                config_key='host_types',
                config_value='物理服务器,虚拟机,云服务器,容器',
                config_type='主机类型',
                description='主机类型选项'
            ),
            SystemConfig(
                config_key='middleware_types',
                config_value='WebLogic,Tomcat,JBoss,IIS,Nginx,Kubernetes,Apache',
                config_type='中间件类型',
                description='中间件类型选项'
            ),
            SystemConfig(
                config_key='database_types',
                config_value='Oracle,MySQL,PostgreSQL,SQL Server,MongoDB,Redis',
                config_type='数据库类型',
                description='数据库类型选项'
            ),
            SystemConfig(
                config_key='system_status',
                config_value='运行中,维护中,已停用',
                config_type='系统状态',
                description='系统状态选项'
            )
        ]
        
        for config in configs:
            db.session.add(config)
        
        # 创建默认管理员账号
        admin = User(
            username='admin',
            real_name='系统管理员',
            role='admin',
            department='运维中心',
            email='admin@example.com'
        )
        admin.set_password('Flzx3qc@2024')
        db.session.add(admin)
        
        # 示例计划任务
        if not PlanTask.query.first():
            sample_task = PlanTask(
                title='月度备份磁带归档',
                task_type='备份类',
                schedule_type='monthly',
                schedule_value='每月1日',
                plan_time=datetime.now() + timedelta(days=3),
                reminder_minutes=1440,
                reminder_enabled=True,
                alert_robot='默认钉钉机器人',
                webhook_url='https://oapi.dingtalk.com/robot/send?access_token=demo',
                reminder_message='【计划任务提醒】任务：{title}\n时间：{plan_time}\n负责人：{owner}\n准备：{preparations}',
                status='待执行',
                responsible='张三,李四',
                owner='张三',
                description='完成上月磁带备份归档并更新登记表',
                created_by='admin'
            )

            db.session.add(sample_task)
            db.session.flush()
            preparations = [
                '检查磁带完整性',
                '准备归档登记表',
                '确认备份日志无异常'
            ]
            for idx, text in enumerate(preparations, start=1):
                db.session.add(PlanTaskPreparation(
                    task_id=sample_task.id,
                    description=text,
                    status='未开始',
                    order_no=idx
                ))
        
        db.session.commit()
        print('数据库初始化完成!')


if __name__ == '__main__':
    init_database()
    app.run(debug=True, host='0.0.0.0', port=5001)

